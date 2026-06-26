"""Shared question import pipeline.

This module centralizes template parsing, validation, preview creation,
error-file generation, and optional DB persistence so both API views and
management commands can use exactly the same rules.
"""

import io
import math
import os
from dataclasses import dataclass
from typing import Any

import pandas as pd
from django.db import transaction
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from assessment.models import (
    ASSESSMENT_LEVELS,
    MAP_SUBLEVELS_TO_SUBJECT,
    SUB_LEVELS,
    Content,
    Option,
    Question,
    Subject,
    User,
)


REQUIRED_HEADERS = [
    "No",
    "Text Question",
    "Option1",
    "Option2",
    "Option3",
    "Option4",
    "Option5",
    "Correct Option",
]
ALLOWED_QUESTION_TYPES = {"MCQ"}


@dataclass
class ImportResult:
    """Container for the complete outcome of an import validation run."""

    is_valid: bool
    message: str
    error_messages: list[str]
    meta_errors: dict[int, str]
    row_errors: dict[int, str]
    meta_df: pd.DataFrame
    data_df: pd.DataFrame
    normalized_meta: dict[str, Any]
    preview_questions: list[dict[str, Any]]
    error_file_bytes: bytes | None = None
    error_file_name: str | None = None


def is_empty_or_nan(value):
    """Return True when the incoming cell value should be treated as empty."""

    if value is None or value == "nan" or (isinstance(value, str) and not value.strip()):
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return False


def normalize_choice(value, code_to_label):
    """Normalize a metadata value to its canonical choice code.

    Accepts either:
    - The raw code (e.g. ``G1L1``), or
    - The display label (e.g. ``Group 1 Level 1``)
    """

    if is_empty_or_nan(value):
        return None

    normalized = str(value).strip()
    if normalized in code_to_label:
        return normalized

    label_to_code = {
        str(label).strip(): code
        for code, label in code_to_label.items()
    }
    return label_to_code.get(normalized)


def _build_frames_from_sheet(sheet_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Split a raw sheet into metadata rows and question rows.

    Expected template layout:
    - Row 1-7: metadata
    - Row 8: question headers
    - Row 9 onward: question data
    """

    meta_df = sheet_df.iloc[:7, :].copy()
    meta_df = meta_df.reindex(columns=range(max(3, meta_df.shape[1])), fill_value="")
    if 2 not in meta_df.columns:
        meta_df[2] = ""
    meta_df = meta_df.iloc[:, :3]

    header_row = sheet_df.iloc[7].fillna("").astype(str).str.strip().tolist() if len(sheet_df) > 7 else []
    if not header_row:
        header_row = REQUIRED_HEADERS[:]

    data_df = sheet_df.iloc[8:, :].copy() if len(sheet_df) > 8 else pd.DataFrame()
    if data_df.empty:
        data_df = pd.DataFrame(columns=header_row)
    else:
        data_df.columns = header_row[: len(data_df.columns)]

    for header in REQUIRED_HEADERS:
        if header not in data_df.columns:
            data_df[header] = ""

    if "Status" not in data_df.columns:
        data_df["Status"] = ""

    return meta_df, data_df


def _load_template_frames(template_source) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load template content from path or uploaded file object.

    Supports CSV and Excel formats. Returned data is normalized using
    ``_build_frames_from_sheet``.
    """

    if hasattr(template_source, "read"):
        raw_bytes = template_source.read()
        if hasattr(template_source, "seek"):
            template_source.seek(0)
        stream = io.BytesIO(raw_bytes)
        name = getattr(template_source, "name", "") or ""
        ext = os.path.splitext(name.lower())[1]
        if ext == ".csv":
            csv_df = pd.read_csv(stream, header=None, dtype=str, keep_default_na=False).fillna("")
            return _build_frames_from_sheet(csv_df)
        sheet_df = pd.read_excel(stream, header=None, engine="openpyxl").fillna("")
        return _build_frames_from_sheet(sheet_df)

    path = str(template_source)
    ext = os.path.splitext(path.lower())[1]
    if ext == ".csv":
        csv_df = pd.read_csv(path, header=None, dtype=str, keep_default_na=False).fillna("")
        return _build_frames_from_sheet(csv_df)
    sheet_df = pd.read_excel(path, header=None, engine="openpyxl").fillna("")
    return _build_frames_from_sheet(sheet_df)


def _build_preview_questions(df: pd.DataFrame) -> list[dict[str, Any]]:
    """Build frontend-ready question payload from validated rows only."""

    preview = []
    for _, row in df.iterrows():
        if row.get("Status") != "Validated":
            continue
        question_text = str(row.get("Text Question", "")).strip()
        options = []
        for i in range(1, 6):
            value = row.get(f"Option{i}", "")
            if not is_empty_or_nan(value):
                options.append(str(value).strip())
        correct = str(row.get("Correct Option", "")).strip()
        correct_index = options.index(correct) + 1 if correct in options else 0
        preview.append(
            {
                "id": "new",
                "question": question_text,
                "options": [{"option": opt} for opt in options],
                "correct": str(correct_index),
                "category": "",
                "type": "question",
                "having_assessment": "False",
            }
        )
    return preview


def _validate_import(meta_df: pd.DataFrame, df: pd.DataFrame) -> tuple[dict[str, Any], dict[int, str], dict[int, str], list[str]]:
    """Validate metadata and all question rows in one pass.

    Returns:
    - normalized metadata
    - metadata error map (metadata row index -> error)
    - question row error map (dataframe index -> error)
    - flattened list of human-readable errors for alert/JSON responses
    """

    level_choices = {code: label for code, label in ASSESSMENT_LEVELS}
    sub_level_choices = {code: label for code, label in SUB_LEVELS}
    subject_choices = {choice.value: choice.label for choice in Subject}

    meta_errors: dict[int, str] = {}
    row_errors: dict[int, str] = {}
    all_errors: list[str] = []

    def add_meta_error(row_index: int, message: str):
        meta_errors[row_index] = message
        meta_df.at[row_index, 2] = message
        all_errors.append(f"Row {row_index + 1}: {message}")

    content = str(meta_df.iloc[0, 1]).strip() if len(meta_df) > 0 else ""
    level = str(meta_df.iloc[1, 1]).strip() if len(meta_df) > 1 else ""
    subject = str(meta_df.iloc[2, 1]).strip() if len(meta_df) > 2 else ""
    sub_level = str(meta_df.iloc[3, 1]).strip() if len(meta_df) > 3 else ""
    question_type = str(meta_df.iloc[4, 1]).strip() if len(meta_df) > 4 else ""
    author_email = str(meta_df.iloc[5, 1]).strip() if len(meta_df) > 5 else ""

    normalized_level = None
    normalized_subject = None
    normalized_sub_level = None
    author = None

    if is_empty_or_nan(content):
        add_meta_error(0, "Content not found")

    if is_empty_or_nan(level):
        add_meta_error(1, "Level not found")
    else:
        normalized_level = normalize_choice(level, level_choices)
        if not normalized_level:
            add_meta_error(1, "Invalid level")

    if is_empty_or_nan(subject):
        add_meta_error(2, "Subject not found")
    else:
        normalized_subject = normalize_choice(subject, subject_choices)
        if not normalized_subject:
            add_meta_error(2, "Invalid subject")

    if is_empty_or_nan(sub_level):
        add_meta_error(3, "Sub-level not found")
    else:
        normalized_sub_level = normalize_choice(sub_level, sub_level_choices)
        if not normalized_sub_level:
            add_meta_error(3, "Invalid sub-level")

    if is_empty_or_nan(question_type):
        add_meta_error(4, "Type not found")
    elif question_type not in ALLOWED_QUESTION_TYPES:
        add_meta_error(4, "Invalid type")

    if normalized_subject and normalized_sub_level:
        subject_sub_levels = {
            code for code, _ in MAP_SUBLEVELS_TO_SUBJECT.get(normalized_subject, [])
        }
        if subject_sub_levels and normalized_sub_level not in subject_sub_levels:
            add_meta_error(3, "Invalid sub-level")

    if is_empty_or_nan(author_email):
        add_meta_error(5, "Author email not found")
    else:
        author = User.objects.filter(email__iexact=author_email).first()
        if not author:
            add_meta_error(5, "Invalid email")

    actual_headers = [str(c).strip() for c in df.columns.tolist()]
    # Header mismatch is recorded once; row-level validation still continues so
    # users can fix all issues in a single attempt.
    if actual_headers[: len(REQUIRED_HEADERS)] != REQUIRED_HEADERS:
        row_errors[-1] = "Import failed (Header mismatch)"
        all_errors.append(
            f"Header mismatch. Expected {REQUIRED_HEADERS}, got {actual_headers[:len(REQUIRED_HEADERS)]}"
        )

    for row_index, row in df.iterrows():
        row_problems: list[str] = []

        question_text = row.get("Text Question")
        options = []
        for i in range(1, 6):
            opt = row.get(f"Option{i}", "")
            if not is_empty_or_nan(opt):
                options.append(str(opt).strip())
        correct = str(row.get("Correct Option", "")).strip()

        is_empty_row = is_empty_or_nan(question_text) and not options and is_empty_or_nan(correct)
        if is_empty_row:
            # Keep completely blank trailing/template rows ignored.
            df.at[row_index, "Status"] = ""
            continue

        if is_empty_or_nan(question_text):
            row_problems.append("Empty question")
        if len(options) < 2:
            row_problems.append("At least 2 options are required")
        if is_empty_or_nan(correct):
            row_problems.append("Missing correct option")
        elif correct not in options:
            row_problems.append(f"Correct option '{correct}' not in options")

        if row_problems:
            joined = "; ".join(row_problems)
            status = f"Import failed ({joined})"
            row_errors[row_index] = status
            df.at[row_index, "Status"] = status
            all_errors.append(f"Question row {row_index + 9}: {joined}")
        else:
            df.at[row_index, "Status"] = "Validated"

    normalized_meta = {
        "content": content,
        "level": normalized_level,
        "subject": normalized_subject,
        "sub_level": normalized_sub_level,
        "question_type": question_type,
        "author": author,
        "author_email": author_email,
    }
    return normalized_meta, meta_errors, row_errors, all_errors


def _write_status_workbook(meta_df: pd.DataFrame, df: pd.DataFrame) -> bytes:
    """Generate Excel bytes containing metadata, row status, and highlighting."""

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        meta_df.to_excel(writer, index=False, header=False)
        df.to_excel(writer, index=False, startrow=len(meta_df) )

        workbook = writer.book
        worksheet = writer.sheets[list(writer.sheets.keys())[0]]

        choices_sheet_name = "MetaChoices"
        if choices_sheet_name in workbook.sheetnames:
            choices_ws = workbook[choices_sheet_name]
        else:
            choices_ws = workbook.create_sheet(title=choices_sheet_name)

        level_codes = [str(code) for code, _ in ASSESSMENT_LEVELS]
        subject_codes = [str(choice.value) for choice in Subject]
        sublevel_codes = [str(code) for code, _ in SUB_LEVELS]

        choices_ws["A1"] = "Levels"
        choices_ws["B1"] = "Subjects"
        choices_ws["C1"] = "SubLevelsAll"

        for idx, value in enumerate(level_codes, start=2):
            choices_ws.cell(row=idx, column=1, value=value)
        for idx, value in enumerate(subject_codes, start=2):
            choices_ws.cell(row=idx, column=2, value=value)
        for idx, value in enumerate(sublevel_codes, start=2):
            choices_ws.cell(row=idx, column=3, value=value)

        start_col = 5
        for idx, subject_code in enumerate(subject_codes):
            col = start_col + idx
            choices_ws.cell(row=1, column=col, value=subject_code)
            mapped = [code for code, _ in MAP_SUBLEVELS_TO_SUBJECT.get(subject_code, [])]
            if not mapped:
                mapped = sublevel_codes[:]
            for row_idx, value in enumerate(mapped, start=2):
                choices_ws.cell(row=row_idx, column=col, value=value)

            col_letter = choices_ws.cell(row=1, column=col).column_letter
            end_row = 1 + len(mapped)
            workbook.defined_names.add(
                DefinedName(
                    name=subject_code,
                    attr_text=f"'{choices_sheet_name}'!${col_letter}$2:${col_letter}${end_row}",
                )
            )

        choices_ws.sheet_state = "hidden"

        level_end = 1 + len(level_codes)
        subject_end = 1 + len(subject_codes)

        level_dv = DataValidation(
            type="list",
            formula1=f"='{choices_sheet_name}'!$A$2:$A${level_end}",
            allow_blank=False,
        )
        subject_dv = DataValidation(
            type="list",
            formula1=f"='{choices_sheet_name}'!$B$2:$B${subject_end}",
            allow_blank=False,
        )
        sublevel_dv = DataValidation(
            type="list",
            formula1="=INDIRECT($B$3)",
            allow_blank=False,
        )

        worksheet.add_data_validation(level_dv)
        worksheet.add_data_validation(subject_dv)
        worksheet.add_data_validation(sublevel_dv)
        level_dv.add("B2")
        subject_dv.add("B3")
        sublevel_dv.add("B4")

        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for meta_row in range(1, len(meta_df) + 1):
            meta_error = worksheet.cell(row=meta_row, column=3).value
            if meta_error:
                for col in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=meta_row, column=col).fill = error_fill

        status_col_idx = list(df.columns).index("Status") + 1
        # Data starts after: metadata rows + blank row + header row.
        first_data_row = len(meta_df) + 3
        for df_index in range(len(df)):
            excel_row = first_data_row + df_index
            status_value = worksheet.cell(row=excel_row, column=status_col_idx).value
            if status_value and str(status_value).startswith("Import failed"):
                for col in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=excel_row, column=col).fill = error_fill

    output.seek(0)
    return output.read()


def build_sample_template_workbook_bytes() -> bytes:
    """Generate a sample import template (with metadata dropdowns) as Excel bytes."""

    sample_rows = [
        ["Content", "Sample content for imported questions.", ""],
        ["Level", "G1L1", ""],
        ["Subject", "EN", ""],
        ["Sub-level", "General", ""],
        ["Question Type", "MCQ", ""],
        ["Author Email", "your-email@example.com", ""],
        ["", "", ""],
    ]
    meta_df = pd.DataFrame(sample_rows)
    data_df = pd.DataFrame(
        [
            ["1", "What is 2 + 2?", "2", "3", "4", "", "", "4", ""],
        ],
        columns=REQUIRED_HEADERS + ["Status"],
    )
    return _write_status_workbook(meta_df, data_df)


def validate_and_prepare_import(template_source, source_name: str = "questions.xlsx") -> ImportResult:
    """Parse + validate a template and return preview data without DB writes.

    This is the entry point used by API views for soft-import/preview.
    """

    meta_df, data_df = _load_template_frames(template_source)
    normalized_meta, meta_errors, row_errors, all_errors = _validate_import(meta_df, data_df)
    is_valid = len(meta_errors) == 0 and len(row_errors) == 0
    preview_questions = _build_preview_questions(data_df)

    error_file_bytes = None
    error_file_name = None
    message = "Import successful"
    if not is_valid:
        message = "Import failed. Please download the error file."
        error_file_bytes = _write_status_workbook(meta_df, data_df)
        base_name = os.path.splitext(os.path.basename(source_name))[0] or "questions"
        error_file_name = f"{base_name}_import_errors.xlsx"

    return ImportResult(
        is_valid=is_valid,
        message=message,
        error_messages=all_errors,
        meta_errors=meta_errors,
        row_errors=row_errors,
        meta_df=meta_df,
        data_df=data_df,
        normalized_meta=normalized_meta,
        preview_questions=preview_questions,
        error_file_bytes=error_file_bytes,
        error_file_name=error_file_name,
    )


def persist_validated_import(import_result: ImportResult) -> None:
    """Persist questions for a previously validated import result."""

    if not import_result.is_valid:
        raise ValueError("Cannot persist invalid import result")

    normalized_meta = import_result.normalized_meta
    with transaction.atomic():
        content_obj = Content.objects.create(content=normalized_meta["content"])
        for row_index, row in import_result.data_df.iterrows():
            if row.get("Status") != "Validated":
                continue

            question_text = str(row.get("Text Question", "")).strip()
            options = []
            for i in range(1, 6):
                opt = row.get(f"Option{i}", "")
                if not is_empty_or_nan(opt):
                    options.append(str(opt).strip())
            correct = str(row.get("Correct Option", "")).strip()

            question_obj = Question.objects.create(
                question=question_text,
                content=content_obj,
                level=normalized_meta["level"],
                sub_level=normalized_meta["sub_level"],
                subject=normalized_meta["subject"],
                author=normalized_meta["author"],
            )

            option_objs = [Option(option=opt, is_correct=(opt == correct)) for opt in options]
            created_opts = Option.objects.bulk_create(option_objs)
            question_obj.options.add(*created_opts)
            # Mutate status so exported output reflects what was inserted.
            import_result.data_df.at[row_index, "Status"] = "Imported successfully"


def write_status_file_to_path(path: str, meta_df: pd.DataFrame, data_df: pd.DataFrame) -> str:
    """Write workbook with status/errors beside source file and return path."""

    file_root, file_ext = os.path.splitext(os.path.basename(path))
    output_ext = file_ext if file_ext.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"} else ".xlsx"
    output_path = os.path.join(os.path.dirname(path), f"{file_root}_imported{output_ext}")
    content = _write_status_workbook(meta_df, data_df)
    with open(output_path, "wb") as fp:
        fp.write(content)
    return output_path
